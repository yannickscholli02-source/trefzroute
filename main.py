"""
TrefzRoute Backend - FastAPI + Google OR-Tools VRP
Zeitfenster: erste Zustellung ab 07:30, letzte bis 16:30 (strikt)
Abfahrt: so spaet wie moeglich (Fahrzeit zum ersten Stopp - 07:30)
Rueckfahrt: keine Zeitbeschraenkung
"""

from fastapi import FastAPI, HTTPException, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
import math
import os
import io

try:
    from ortools.constraint_solver import routing_enums_pb2
    from ortools.constraint_solver import pywrapcp
    ORTOOLS_OK = True
    print("[OK] OR-Tools geladen")
except ImportError as e:
    ORTOOLS_OK = False
    print("[WARNUNG] OR-Tools nicht verfuegbar:", e)

app = FastAPI(title="TrefzRoute API", version="3.1.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# -- ZEITKONSTANTEN --
EARLIEST_DELIVERY  = 7 * 3600 + 30 * 60   # 07:30 in Sekunden ab Mitternacht
LATEST_DELIVERY    = 16 * 3600 + 30 * 60   # 16:30 in Sekunden ab Mitternacht
EARLIEST_DEPARTURE = 6 * 3600              # 06:00 früheste Abfahrt
# Depot-Zeitfenster: [0, LATEST_DELIVERY + 3h] (Rückkehr bis 19:30 erlaubt)
LATEST_RETURN      = LATEST_DELIVERY + 3 * 3600  # 19:30

# -- MODELS --

class MatrixOptimizeRequest(BaseModel):
    addresses: List[str]
    dist_matrix: List[List[float]]
    dur_matrix: List[List[float]]          # reine Fahrzeiten
    travel_dur_matrix: Optional[List[List[float]]] = None  # alias, rückwärtskompatibel
    service_times: Optional[List[int]] = None  # Abladezeit pro Knoten in Sekunden (index 0 = depot = 0)
    max_stops_per_route: int = 999
    max_duration_s: Optional[float] = None

class StopResult(BaseModel):
    address: str
    leg_dist_m: float
    leg_dist_text: str
    leg_dur_s: float
    leg_dur_text: str
    arrival_time: Optional[str] = None  # geschätzte Ankunftszeit

class RouteResult(BaseModel):
    id: int
    stops: List[StopResult]
    total_dist_m: float
    total_dist_text: str
    total_dur_s: float
    total_dur_text: str
    departure_time: Optional[str] = None
    return_time: Optional[str] = None
    feasible: bool = True

class OptimizeResponse(BaseModel):
    routes: List[RouteResult]
    total_dist_m: float
    total_dur_s: float
    num_stops: int
    solver: str
    warnings: List[str] = []

# -- HELPERS --

def format_dist(m: float) -> str:
    return f"{round(m)} m" if m < 1000 else f"{m/1000:.1f} km"

def format_dur(s: float) -> str:
    h, m = int(s // 3600), round((s % 3600) / 60)
    return f"{h}h {m}min" if h > 0 else f"{m} min"

def format_time(s: float) -> str:
    """Sekunden ab Mitternacht -> HH:MM"""
    h = int(s // 3600) % 24
    m = int((s % 3600) // 60)
    return f"{h:02d}:{m:02d}"

# -- GREEDY FALLBACK mit Zeitfenstern --

def solve_greedy(dist_matrix, dur_matrix, depot_idx=0):
    n = len(dist_matrix)
    unvisited = list(range(1, n))
    routes = []

    while unvisited:
        route = []
        current = depot_idx
        # Berechne Abfahrtszeit: so dass erste Zustellung um 07:30
        # Wir starten mit dem nächsten Stopp vom Depot und berechnen rückwärts
        current_time = EARLIEST_DEPARTURE  # starte mit 06:00 als Basis

        while unvisited:
            # Finde nächsten erreichbaren Stopp (Ankunft <= 16:30)
            candidates = []
            for s in unvisited:
                travel = dur_matrix[current][s]
                arrival = max(current_time + travel, EARLIEST_DELIVERY)
                if arrival <= LATEST_DELIVERY:
                    candidates.append((s, arrival))

            if not candidates:
                break

            # Nearest neighbor unter den erreichbaren
            nearest, arrival = min(candidates, key=lambda x: dist_matrix[current][x[0]])
            route.append((nearest, arrival))
            current_time = arrival
            unvisited.remove(nearest)
            current = nearest

        if route:
            routes.append(route)

    return routes

# -- OR-TOOLS VRP MIT ZEITFENSTERN --

def solve_ortools_vrp(dist_matrix, dur_matrix, service_times=None, depot_idx=0):
    n = len(dist_matrix)
    num_stops = n - 1

    num_vehicles = min(num_stops, max(4, math.ceil(num_stops / 8)))
    print(f"[INFO] {num_stops} Stopps, {num_vehicles} max. Fahrzeuge, Zeitfenster 07:30-16:30")

    idist = [[int(dist_matrix[i][j]) for j in range(n)] for i in range(n)]
    idur  = [[int(dur_matrix[i][j])  for j in range(n)] for i in range(n)]

    # Service time pro Knoten (Abladezeit) — wird zur Transitzeit addiert
    isvc = service_times if service_times else [0] * n

    manager = pywrapcp.RoutingIndexManager(n, num_vehicles, depot_idx)
    routing = pywrapcp.RoutingModel(manager)

    def dist_cb(fi, ti):
        return idist[manager.IndexToNode(fi)][manager.IndexToNode(ti)]
    dist_ci = routing.RegisterTransitCallback(dist_cb)
    routing.SetArcCostEvaluatorOfAllVehicles(dist_ci)

    # Zeit-Callback: Fahrzeit VON fi + Abladezeit AM Knoten fi (nicht ti!)
    # So wird die Abladezeit korrekt dem Knoten zugeordnet von dem gefahren wird
    def time_cb(fi, ti):
        from_node = manager.IndexToNode(fi)
        return idur[from_node][manager.IndexToNode(ti)] + isvc[from_node]
    time_ci = routing.RegisterTransitCallback(time_cb)

    # Zeit-Dimension mit Zeitfenstern
    # Horizon: von 06:00 (Abfahrt) bis 19:30 (späteste Rückkehr)
    horizon = LATEST_RETURN
    # max_slack = horizon: keine Begrenzung der Wartezeit — OR-Tools kann
    # Fahrzeuge flexibel über den ganzen Tag verteilen
    routing.AddDimension(time_ci, horizon, horizon, False, "Time")
    time_dim = routing.GetDimensionOrDie("Time")

    # Zeitfenster für jeden Stopp: 07:30 – 16:30
    for i in range(1, n):  # alle außer Depot
        idx = manager.NodeToIndex(i)
        time_dim.CumulVar(idx).SetRange(EARLIEST_DELIVERY, LATEST_DELIVERY)

    # Depot: Abfahrt frühestens 06:00, Rückkehr spätestens 19:30
    for v in range(num_vehicles):
        start_idx = routing.Start(v)
        end_idx   = routing.End(v)
        time_dim.CumulVar(start_idx).SetRange(EARLIEST_DEPARTURE, LATEST_DELIVERY)
        time_dim.CumulVar(end_idx).SetRange(0, LATEST_RETURN)

    # Minimiere Gesamtdistanz (primary) und Fahrzeuganzahl (secondary)
    routing.SetFixedCostOfAllVehicles(100_000)

    # Suchparameter — 8s reichen für ~35 Stopps pro Cluster
    params = pywrapcp.DefaultRoutingSearchParameters()
    params.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
    params.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    params.time_limit.seconds = 8
    params.log_search = False

    sol = routing.SolveWithParameters(params)

    if not sol:
        # Schneller Retry mit anderer Strategie
        params.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.SAVINGS
        params.time_limit.seconds = 5
        sol = routing.SolveWithParameters(params)

    if not sol:
        return None, manager, routing, None

    return sol, manager, routing, time_dim

# -- ENDPOINTS --

@app.get("/health")
def health():
    return {
        "status": "ok",
        "solver": "OR-Tools VRP + Zeitfenster" if ORTOOLS_OK else "Greedy",
        "ortools_available": ORTOOLS_OK,
        "version": "3.1.0"
    }

@app.post("/optimize_matrix", response_model=OptimizeResponse)
def optimize_matrix(req: MatrixOptimizeRequest):
    n = len(req.addresses)
    if n < 2:
        raise HTTPException(400, "Mindestens 1 Stopp erforderlich.")

    dist_matrix  = req.dist_matrix
    dur_matrix   = req.dur_matrix
    service_times = req.service_times  # Abladezeit pro Knoten
    # Für Anzeige: reine Fahrzeit (dur_matrix ist jetzt bereits reine Fahrzeit)
    travel_matrix = req.dur_matrix
    results, total_dist, total_dur, warnings = [], 0, 0, []

    if ORTOOLS_OK:
        try:
            sol, manager, routing, time_dim = solve_ortools_vrp(dist_matrix, dur_matrix, service_times)

            if sol is None:
                # Fallback auf Greedy
                raise Exception("Keine Loesung gefunden")

            solver_name = "OR-Tools VRP + Zeitfenster 07:30-16:30"
            num_vehicles = routing.vehicles()

            for v in range(num_vehicles):
                route_indices = []
                arrivals = []
                idx = routing.Start(v)
                while not routing.IsEnd(idx):
                    node = manager.IndexToNode(idx)
                    if node != 0:
                        arrival_s = sol.Value(time_dim.CumulVar(idx))
                        route_indices.append(node)
                        arrivals.append(arrival_s)
                    idx = sol.Value(routing.NextVar(idx))

                if not route_indices:
                    continue

                # Abfahrtszeit: rückwärts vom ersten Stopp — nur reine Fahrzeit
                travel_to_first = int(dur_matrix[0][route_indices[0]])
                departure_s = max(EARLIEST_DEPARTURE, arrivals[0] - travel_to_first)

                stops_out, route_dist, route_dur, prev = [], 0, 0, 0
                for i, node in enumerate(route_indices):
                    leg_dist_m   = dist_matrix[prev][node]
                    leg_travel_s = travel_matrix[prev][node]  # reine Fahrzeit für Anzeige
                    leg_dur_s    = dur_matrix[prev][node]     # inkl. Abladezeit für Gesamtdauer
                    route_dist += leg_dist_m
                    route_dur  += leg_dur_s
                    stops_out.append(StopResult(
                        address=req.addresses[node],
                        leg_dist_m=leg_dist_m,
                        leg_dist_text=format_dist(leg_dist_m),
                        leg_dur_s=leg_travel_s,
                        leg_dur_text=format_dur(leg_travel_s),
                        arrival_time=format_time(arrivals[i]),
                    ))
                    prev = node

                # Rückfahrt: letzte Ankunftszeit + Fahrzeit zurück zum Depot
                last_arrival_s = arrivals[-1] if arrivals else departure_s
                return_dist_m = dist_matrix[prev][0]
                return_dur_s  = dur_matrix[prev][0]
                # Rückkehr-Uhrzeit = letzte Ankunft + Abladezeit bereits in dur enthalten
                # dur_matrix[prev][0] enthält keine Abladezeit (j=0 ist Depot)
                return_time_s = last_arrival_s + return_dur_s

                route_dist += return_dist_m
                route_dur  += return_dur_s
                total_dist += route_dist
                total_dur  += route_dur

                # Prüfe ob letzte Zustellung vor 16:30
                last_arrival = arrivals[-1] if arrivals else 0
                feasible = last_arrival <= LATEST_DELIVERY
                if not feasible:
                    warnings.append(f"Route {len(results)+1}: letzte Zustellung {format_time(last_arrival)} > 16:30")

                results.append(RouteResult(
                    id=len(results) + 1,
                    stops=stops_out,
                    total_dist_m=route_dist,
                    total_dist_text=format_dist(route_dist),
                    total_dur_s=route_dur,
                    total_dur_text=format_dur(route_dur),
                    departure_time=format_time(departure_s),
                    return_time=format_time(return_time_s),
                    feasible=feasible,
                ))

        except Exception as e:
            print(f"OR-Tools Fehler: {e} — Greedy Fallback")
            greedy_routes = solve_greedy(dist_matrix, dur_matrix)
            solver_name = "Greedy + Zeitfenster (Fallback)"

            for ri, route in enumerate(greedy_routes):
                stops_out, route_dist, route_dur, prev = [], 0, 0, 0
                departure_s = None
                last_arrival_s = 0
                for node, arrival_s in route:
                    leg_dist_m   = dist_matrix[prev][node]
                    leg_travel_s = travel_matrix[prev][node]
                    leg_dur_s    = dur_matrix[prev][node]
                    if departure_s is None:
                        departure_s = max(EARLIEST_DEPARTURE, arrival_s - leg_travel_s)
                    route_dist += leg_dist_m
                    route_dur  += leg_dur_s
                    last_arrival_s = arrival_s
                    stops_out.append(StopResult(
                        address=req.addresses[node],
                        leg_dist_m=leg_dist_m,
                        leg_dist_text=format_dist(leg_dist_m),
                        leg_dur_s=leg_travel_s,
                        leg_dur_text=format_dur(leg_travel_s),
                        arrival_time=format_time(arrival_s),
                    ))
                    prev = node
                return_dur_s = dur_matrix[prev][0]
                return_time_s = last_arrival_s + return_dur_s
                route_dist += dist_matrix[prev][0]
                route_dur  += return_dur_s
                total_dist += route_dist
                total_dur  += route_dur
                results.append(RouteResult(
                    id=ri + 1, stops=stops_out,
                    total_dist_m=route_dist, total_dist_text=format_dist(route_dist),
                    total_dur_s=route_dur,  total_dur_text=format_dur(route_dur),
                    departure_time=format_time(departure_s) if departure_s else None,
                    return_time=format_time(return_time_s),
                    feasible=True,
                ))
    else:
        greedy_routes = solve_greedy(dist_matrix, dur_matrix)
        solver_name = "Greedy + Zeitfenster"
        for ri, route in enumerate(greedy_routes):
            stops_out, route_dist, route_dur, prev = [], 0, 0, 0
            for node, arrival_s in route:
                leg_dist_m   = dist_matrix[prev][node]
                leg_travel_s = travel_matrix[prev][node]
                leg_dur_s    = dur_matrix[prev][node]
                route_dist += leg_dist_m
                route_dur  += leg_dur_s
                stops_out.append(StopResult(
                    address=req.addresses[node],
                    leg_dist_m=leg_dist_m, leg_dist_text=format_dist(leg_dist_m),
                    leg_dur_s=leg_travel_s, leg_dur_text=format_dur(leg_travel_s),
                    arrival_time=format_time(arrival_s),
                ))
                prev = node
            route_dist += dist_matrix[prev][0]
            route_dur  += dur_matrix[prev][0]
            total_dist += route_dist
            total_dur  += route_dur
            results.append(RouteResult(
                id=ri + 1, stops=stops_out,
                total_dist_m=route_dist, total_dist_text=format_dist(route_dist),
                total_dur_s=route_dur,  total_dur_text=format_dur(route_dur),
                feasible=True,
            ))

    return OptimizeResponse(
        routes=results, total_dist_m=total_dist, total_dur_s=total_dur,
        num_stops=n - 1, solver=solver_name, warnings=warnings,
    )

# -- EXCEL EXPORT --

class ExcelStop(BaseModel):
    addr: str
    arrivalTime: Optional[str] = None
    legDistText: str
    legDurText: str
    legDistM: float = 0
    schulname: Optional[str] = ''
    tel: Optional[str] = ''
    email: Optional[str] = ''

class ExcelRoute(BaseModel):
    id: int
    stops: List[ExcelStop]
    totalDist: float
    totalDur: float
    departureTime: Optional[str] = None
    returnTime: Optional[str] = None

class ExcelExportRequest(BaseModel):
    routes: List[ExcelRoute]
    depot: str
    single_route_id: Optional[int] = None
    has_grunddaten: bool = False

def round_to_five(minutes: int) -> int:
    return round(minutes / 5) * 5

def calc_zeitfenster(arrival_time: Optional[str]) -> str:
    if not arrival_time or arrival_time == '—':
        return '—'
    hh, mm = map(int, arrival_time.split(':'))
    total = hh * 60 + mm
    rounded = round_to_five(total)
    from_min = max(7 * 60, rounded - 60)
    to_min = rounded + 60
    def fmt(m): return f"{m//60:02d}:{m%60:02d}"
    return f"{fmt(from_min)} – {fmt(to_min)}"

def real_duration_min(r: ExcelRoute) -> int:
    if r.departureTime and r.returnTime:
        dh, dm = map(int, r.departureTime.split(':'))
        rh, rm = map(int, r.returnTime.split(':'))
        return (rh * 60 + rm) - (dh * 60 + dm)
    return round(r.totalDur / 60)

def build_route_sheet(wb, r: ExcelRoute, depot: str, has_grunddaten: bool = False):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_BG  = '305496'
    DEPOT_BG   = 'FFE699'
    BORDER_CLR = 'BFBFBF'

    thin = Side(style='thin', color=BORDER_CLR)
    def brd(): return Border(left=thin, right=thin, top=thin, bottom=thin)
    def hdr_font(): return Font(name='Arial', size=10, bold=True, color='FFFFFF')
    def hdr_fill(): return PatternFill('solid', fgColor=HEADER_BG)
    def depot_font(): return Font(name='Arial', size=10, bold=True)
    def depot_fill(): return PatternFill('solid', fgColor=DEPOT_BG)
    def norm_font(): return Font(name='Arial', size=10)
    def wrap_align(center=False):
        return Alignment(wrap_text=True, vertical='center', horizontal='center' if center else 'left')

    ws = wb.create_sheet(title=f'Route {r.id}')

    km = f"{r.totalDist/1000:.1f}"
    dur_min = real_duration_min(r)
    h, m = divmod(dur_min, 60)
    dep = r.departureTime or '—'
    last_arr = r.stops[-1].arrivalTime if r.stops else '—'
    ret = r.returnTime or '—'

    # Spalten definieren
    if has_grunddaten:
        headers = ['Stopp-Nr', 'Schulname', 'Adresse', 'Telefon', 'E-Mail', 'Zeitfenster', 'Strecke zum Stopp', 'Fahrzeit zum Stopp', 'Abladezeit']
        COLS = 9
    else:
        headers = ['Stopp-Nr', 'Adresse', 'Zeitfenster', 'Strecke zum Stopp', 'Fahrzeit zum Stopp', 'Abladezeit']
        COLS = 6

    col_letters = [get_column_letter(i+1) for i in range(COLS)]

    # Titelzeilen
    ws['A1'] = f'TrefzRoute — Route {r.id}'
    ws['A1'].font = Font(name='Arial', size=10, bold=True)
    ws['A2'] = f'Datum: {__import__("datetime").date.today().strftime("%d.%m.%Y")}   |   Depot: {depot}'
    ws['A2'].font = norm_font()
    ws['A3'] = f'Gesamt: {len(r.stops)} Stopps   |   {km} km   |   {h}h {m}min   |   Abfahrt: {dep}   |   Letzte Zustellung: {last_arr}   |   Rückkehr: {ret}'
    ws['A3'].font = norm_font()
    for row in [1, 2, 3]:
        ws.merge_cells(f'A{row}:{col_letters[-1]}{row}')

    # Header
    for ci, h_txt in enumerate(headers, 1):
        cell = ws.cell(row=5, column=ci, value=h_txt)
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.border = brd()
        cell.alignment = wrap_align(center=(ci == 1))

    # Abfahrt
    if has_grunddaten:
        depot_row_data = ['Abfahrt', '', depot, '', '', dep, '—', '—', '—']
    else:
        depot_row_data = ['Abfahrt', depot, dep, '—', '—', '—']
    for ci, val in enumerate(depot_row_data, 1):
        cell = ws.cell(row=6, column=ci, value=val)
        cell.font = depot_font()
        cell.fill = depot_fill()
        cell.border = brd()
        cell.alignment = wrap_align(center=(ci == 1))

    # Stopps
    addr_seen = {}
    for i, s in enumerate(r.stops):
        row_num = 7 + i
        is_dup = s.legDistM == 0 or s.legDistText == '0 m'
        svc = '5 min' if s.addr in addr_seen else '15 min'
        addr_seen[s.addr] = True
        if has_grunddaten:
            row_data = [
                f'{r.id}/{i+1}',
                s.schulname or '',
                s.addr,
                s.tel or '',
                s.email or '',
                calc_zeitfenster(s.arrivalTime),
                '0 m' if is_dup else s.legDistText,
                '0 min' if is_dup else s.legDurText,
                svc,
            ]
        else:
            row_data = [
                f'{r.id}/{i+1}',
                s.addr,
                calc_zeitfenster(s.arrivalTime),
                '0 m' if is_dup else s.legDistText,
                '0 min' if is_dup else s.legDurText,
                svc,
            ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.font = norm_font()
            cell.border = brd()
            cell.alignment = wrap_align(center=(ci == 1))

    # Rückkehr
    ret_row_num = 7 + len(r.stops)
    if has_grunddaten:
        ret_data = ['Rückkehr', '', depot, '', '', ret, '—', '—', '—']
    else:
        ret_data = ['Rückkehr', depot, ret, '—', '—', '—']
    for ci, val in enumerate(ret_data, 1):
        cell = ws.cell(row=ret_row_num, column=ci, value=val)
        cell.font = depot_font()
        cell.fill = depot_fill()
        cell.border = brd()
        cell.alignment = wrap_align(center=(ci == 1))

    # Spaltenbreiten
    addr_max = max((len(s.addr) for s in r.stops), default=20)
    addr_max = max(addr_max, len(depot))
    if has_grunddaten:
        schulname_max = max((len(s.schulname or '') for s in r.stops), default=20)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = min(max(schulname_max * 0.85, 25), 60)
        ws.column_dimensions['C'].width = min(max(addr_max * 0.85, 25), 55)
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 28
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 12
    else:
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = min(max(addr_max * 0.85, 25), 60)
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12

    for row_num in range(5, 7 + len(r.stops) + 1):
        ws.row_dimensions[row_num].height = 30

@app.post('/export_excel')
async def export_excel(req: ExcelExportRequest):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail='openpyxl nicht installiert')

    wb = Workbook()
    wb.remove(wb.active)  # leeres default sheet entfernen

    routes_to_export = req.routes
    if req.single_route_id is not None:
        routes_to_export = [r for r in req.routes if r.id == req.single_route_id]

    if not routes_to_export:
        raise HTTPException(status_code=404, detail='Route nicht gefunden')

    # Übersicht nur bei Gesamtexport
    if req.single_route_id is None:
        from openpyxl.styles import Side
        thin = Side(style='thin', color='BFBFBF')
        def brd(): return Border(left=thin, right=thin, top=thin, bottom=thin)

        ws_sum = wb.create_sheet(title='Übersicht')
        ws_sum['A1'] = 'TrefzRoute — Übersicht aller Routen'
        ws_sum['A1'].font = Font(name='Arial', size=10, bold=True)
        ws_sum.merge_cells('A1:G1')
        ws_sum['A2'] = f'Datum: {__import__("datetime").date.today().strftime("%d.%m.%Y")}   |   Depot: {req.depot}'
        ws_sum.merge_cells('A2:G2')

        headers = ['Route', 'Anzahl Stopps', 'Gesamtstrecke', 'Gesamtdauer', 'Abfahrt', 'Letzte Zustellung', 'Rückkehr']
        for ci, h in enumerate(headers, 1):
            cell = ws_sum.cell(row=4, column=ci, value=h)
            cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='305496')
            cell.border = brd()
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        for i, r in enumerate(req.routes):
            dur_min = real_duration_min(r)
            h2, m2 = divmod(dur_min, 60)
            last_arr = r.stops[-1].arrivalTime if r.stops else '—'
            row_data = [f'Route {r.id}', len(r.stops), f'{r.totalDist/1000:.1f} km',
                        f'{h2}h {m2}min', r.departureTime or '—', last_arr, r.returnTime or '—']
            for ci, val in enumerate(row_data, 1):
                cell = ws_sum.cell(row=5+i, column=ci, value=val)
                cell.font = Font(name='Arial', size=10)
                cell.border = brd()
                cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Gesamt-Zeile
        total_row = 5 + len(req.routes)
        total_km = sum(r.totalDist for r in req.routes) / 1000
        total_stops = sum(len(r.stops) for r in req.routes)
        total_min = sum(real_duration_min(r) for r in req.routes)
        th, tm = divmod(total_min, 60)
        total_data = ['Gesamt', total_stops, f'{total_km:.1f} km', f'{th}h {tm}min', '', '', '']
        for ci, val in enumerate(total_data, 1):
            cell = ws_sum.cell(row=total_row, column=ci, value=val)
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.border = brd()
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        for col in ['A','B','C','D','E','F','G']:
            ws_sum.column_dimensions[col].width = [14,16,16,14,12,20,12]['ABCDEFG'.index(col)]

    for r in routes_to_export:
        build_route_sheet(wb, r, req.depot, req.has_grunddaten)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (f'TrefzRoute_R{str(req.single_route_id).zfill(2)}_{__import__("datetime").date.today()}.xlsx'
                if req.single_route_id else f'TrefzRoute_Alle_{__import__("datetime").date.today()}.xlsx')

    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

class RouteLookupEntry(BaseModel):
    routeId: int
    stopNum: int
    zfFrom: str = ''
    zfTo: str = ''

class RoutenlisteRequest(BaseModel):
    file_base64: str
    filename: str
    route_lookup: dict  # normAddr -> RouteLookupEntry dict
    depot: str = ''

@app.post('/export_routenliste')
async def export_routenliste(
    file: UploadFile,
    route_lookup_json: str = Form(...),
    depot: str = Form(''),
):
    try:
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import json
    except ImportError as e:
        raise HTTPException(status_code=500, detail=str(e))

    route_lookup = json.loads(route_lookup_json)
    raw = await file.read()
    filename = file.filename or 'routenliste.xls'

    if filename.lower().endswith('.xls') and not filename.lower().endswith('.xlsx'):
        try:
            import xlrd
            xls_wb = xlrd.open_workbook(file_contents=raw)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for si in range(xls_wb.nsheets):
                xls_ws = xls_wb.sheet_by_index(si)
                out_ws = wb.create_sheet(title=xls_ws.name)
                for ri in range(xls_ws.nrows):
                    for ci in range(xls_ws.ncols):
                        out_ws.cell(row=ri+1, column=ci+1, value=xls_ws.cell_value(ri, ci))
        except Exception as e:
            raise HTTPException(status_code=500, detail=f'XLS-Konvertierung fehlgeschlagen: {e}')
    else:
        wb = load_workbook(io.BytesIO(raw))

    # Work on first sheet only for column insertion
    ws = wb.worksheets[0]

    # New columns to insert after column A (index 1):
    # Route, Lfd. Nr., Zusatz, Zustellung, Uhrzeit 1, Uhrzeit 2, Fahrer
    NEW_COLS = ['Route', 'Lfd. Nr.', 'Zusatz', 'Zustellung', 'Uhrzeit 1', 'Uhrzeit 2', 'Fahrer']
    N = len(NEW_COLS)  # 7

    # Insert N columns after column A (column B becomes column B+N)
    ws.insert_cols(2, N)

    # Styling
    HEADER_BG = '305496'
    BORDER_CLR = 'BFBFBF'
    thin = Side(style='thin', color=BORDER_CLR)
    def brd(): return Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write headers in row 1 for new columns
    for i, col_name in enumerate(NEW_COLS):
        cell = ws.cell(row=1, column=2+i, value=col_name)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=HEADER_BG)
        cell.border = brd()
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # Also style the existing A1 header to match
    a1 = ws.cell(row=1, column=1)
    a1.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    a1.fill = PatternFill('solid', fgColor=HEADER_BG)

    def norm_addr(s):
        import re
        return re.sub(r'\s+', ' ', str(s).lower().strip())

    def build_addr_key(strasse, plz, ort):
        plz_str = str(int(float(plz))) if plz else ''
        return norm_addr(f'{strasse}, {plz_str} {ort}')

    # Find column indices for Straße, PLZ, Ort in original data
    # After insertion, original col B (idx 1) is now col B+N (idx 1+N = 8)
    # Original headers were: A=Nr, B=RP, C=Schulname... G=Straße, H=PLZ, I=Ort
    # After inserting 7 cols after A: A=Nr, B-H=new, I=RP, J=Schulname..., M=Straße, N=PLZ, O=Ort
    # Original col index (0-based): Straße=6, PLZ=7, Ort=8
    # After insert (1-based openpyxl): Straße = 6+1+N = 14, PLZ=15, Ort=16
    COL_STRASSE = 6 + 1 + N  # 14
    COL_PLZ     = 7 + 1 + N  # 15
    COL_ORT     = 8 + 1 + N  # 16

    matched = 0
    for ri in range(2, ws.max_row + 1):
        strasse = ws.cell(row=ri, column=COL_STRASSE).value or ''
        plz     = ws.cell(row=ri, column=COL_PLZ).value or ''
        ort     = ws.cell(row=ri, column=COL_ORT).value or ''
        if not strasse:
            continue

        key = build_addr_key(strasse, plz, ort)
        entry = route_lookup.get(key)

        if entry:
            matched += 1
            route_id = entry.get('routeId', '')
            stop_num = entry.get('stopNum', '')
            zf_from  = entry.get('zfFrom', '')
            zf_to    = entry.get('zfTo', '')

            vals = [route_id, stop_num, '', '', zf_from, zf_to, '']
            for i, val in enumerate(vals):
                cell = ws.cell(row=ri, column=2+i, value=val)
                cell.font = Font(name='Arial', size=10)
                cell.border = brd()
                cell.alignment = Alignment(vertical='center', horizontal='center' if i < 2 else 'left')
        else:
            # Empty cells with border
            for i in range(N):
                cell = ws.cell(row=ri, column=2+i, value='')
                cell.font = Font(name='Arial', size=10)
                cell.border = brd()

    # Column widths for new cols
    widths = [10, 10, 10, 14, 12, 12, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2+i)].width = w

    print(f'[INFO] Routenliste: {matched} von {ws.max_row-1} Zeilen gematcht')

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    out_name = req.filename.replace('.xls', '_Routen.xlsx').replace('.XLS', '_Routen.xlsx')
    return StreamingResponse(
        out_buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{out_name}"'}
    )

# -- STATIC FRONTEND --
frontend_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.exists(frontend_path):
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
